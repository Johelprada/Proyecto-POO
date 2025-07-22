import pickle
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime #Unicamente me interesa el datetime(class) no el modulo completo
#cambios importantes:
# 1. dejé algunas cosas que estaban en algunos lugares privados y otros publicos o bien todas en publicas o todas privadas
#esto para poder entenderme mejor
#2. en la linea "catalogo_productos[nombre]=Producto" cambie producto por mercancia, no me sirvio hasta q no lo hice 
#3. Añadí la clase que me convierte los datos a el excel, aun no le eh puesto los filtros.
#4. luego hago esto interactivo, de momento cree los productos yo directamente, era para probar
ALMACENES= { "Almacen_principal":[], "Almacen_2":[], "Almacen_3":[] }

catalogo_productos={}

lista_cambios = [] #Lista donde se guardaran los cambios generados por funciones def y metodos de clase

registros_a = {} #registros globales! y para descargar

#Tipos de informe A:Generico , U:Importante P:De prueba
class Registro:
    #TIPOS_R={"P":"Prueba","A":"General", "U":"Urgente" "X":"Trivial"} pruebas experimentales de id Registro
    N_registro=0
    def __init__(self, notas_operador: str, cambios: list = None, tipo:str="P"):
        self.fecha = datetime.now()  # Fecha y hora del registro
        Registro.N_registro += 1
        self.tipo = tipo
        self.id = f"{self.fecha.strftime('%m-%d-%H')}:{self.tipo}{Registro.N_registro:04}" #ID unico cada fecha
        self._notas_operador = notas_operador
        self.cambios = cambios if cambios else []  # Cambios realizados en listas u objetos
        registros_a[self.id] = self #clave para guardar los registros que se crean

    def mostrar_registro(self):
        print("=== Registro ===")
        print(f"ID de registro: {self.id}")
        print(f"Fecha: {self.fecha.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Notas del operador: {self._notas_operador}")
        print("Cambios registrados:")
        if len(self.cambios) == 0:
            print("No hubo ningun cambio")
        else:
            for cambio in self.cambios:
                print(f"-{cambio}\n")
        print("================\n")

    def agregar_nota(self, nueva_nota: str):
        self._notas_operador += f"\n{nueva_nota}"

    def agregar_cambio(self, cambio: str):
        self.cambios.append(cambio)

    def __str__(self):
        return f'{self.id}: Usar "metodo mostrar_registro" para leer'

# Metodos para registros
def descargar():
    nombre_archivo = input("Que nombre le quieres poner a estos registros?: ")
    with open(f"{nombre_archivo}.pkl", "wb") as archivo:
        pickle.dump(registros_a, archivo)
def cargar(nombre_archivo="registros_p.pkl"):
    global registros_a
    with open(nombre_archivo, "rb") as archivo:
        registros_a = pickle.load(archivo)

def crear_registro(notas_op:str, tip:str)->"Registro": #Funcion escencial para generar registros
    New_R=Registro(notas_op, lista_cambios, tip)
    return New_R


class Producto:
    definition="Soy un elemento que clasificara con caracteristicas especificas"
    def __init__(self, nombre:str, precio:float, peso:float, categoria:str, stock:int):
        self._nombre = nombre
        self._precio = precio
        self._contenido = nombre
        self.precio = precio
        self.peso = peso
        self.categoria = categoria
        self.stock = stock
    def reavastecer(self, tantos:int):
        self.stock+=tantos

    def get_peso(self)->float:
        return self.peso

    def get_precio(self)->float:
        return self._precio
    
    def get_objeto(self)->str:
        return self._nombre
    
    def es_fragil(self)->bool:
        return self.categoria == "F"
    
    def __str__(self)->str:
        return f"{self._contenido} con valor de {self._precio}$ por unidad"
  
guia_de_ids={
    "P":"es un paquete de prueba", 
    "F":"el paquete es Fragil" ,
    "X":"el paquete no fue marcado",
    "E":"el paquete es electrico",
    "VA":"el paquete posee una forma de vida"
    }

class Paquete:
    IDs_gen= {"P":0, "F":0 , "X":0, "E":0, "VA":0} #Diccionario permite escalabilidad de seriales!
    lista_de_cambios=[]
    def __init__(self, cantidad:int, contenido:"Producto"):
        #Serie de ID--------------------
        self.clasificado=contenido.categoria
        serial=Paquete.IDs_gen[self.clasificado]
        self.id= f"#{serial:04}{self.clasificado}"
        Paquete.IDs_gen[self.clasificado]= 0 if serial > 9999 else serial + 1
        #-------------------------------
        self.contenido=contenido
        self.cantidad=cantidad
        self._precio=contenido.get_precio()*self.cantidad
        self.peso=contenido.get_peso()*self.cantidad

    def esta_en_almacen(self, almacen:str)->bool:
        if self in ALMACENES[almacen]:
            return True
        else:
            return False
        
    def ver_contenido(self):
        lista_cambios.append(f"Se reviso el contenido del paquete {self.id}")
        return self.contenido._nombre
        
    def guardar_paquete(self, almacen:str):
        if self not in ALMACENES[almacen]:
            ALMACENES[almacen].append(self)
            lista_cambios.append(f"Se guardo el paquete {self.id} en el almacen {almacen}")
        else:
            print("El paquete ya esta en este almacen")

    def guardas_muchos_paquetes(self, almacen:list):
        for elemento in almacen:
            if elemento not in almacen:
                almacen.append(elemento)
            else:
                print(f"El paquete {elemento} ya este en el almacen")
        
    def es_fragil(self):
        return self.contenido.es_fragil()
    
    def __str__(self):
        return f"{self.id}"  #f"Paquete de {self.cantidad} x {self.contenido.contenido} ({'frágil' if self.contenido.es_fragil else 'no frágil'})"

#clase con la q generaremos y usaremos los archivos csv
class ManejadorCSV:
    def __init__(self):
        self.archivo = 'productos_almacen.csv'
    
    def generar_csv_de_almacen(self, nombre_almacen="Almacen_principal"):
        datos = [['ID_Paquete', 'Contenido', 'Cantidad', 'Precio_Unitario', 'Precio_Total', 'Categoria', 'Stock']]
        
        if nombre_almacen in ALMACENES:
            for paquete in ALMACENES[nombre_almacen]:
                fila = [
                    paquete.id,
                    paquete.contenido._nombre,
                    paquete.cantidad,
                    paquete.contenido.get_precio(),
                    paquete._precio,
                    paquete.contenido.categoria,
                    paquete.contenido.stock
                ]
                datos.append(fila)
        
        with open(self.archivo, 'w', newline='') as archivo_csv:
            escritor = csv.writer(archivo_csv)
            escritor.writerows(datos)
        
        print(f"Archivo CSV creado: {self.archivo}")
        lista_cambios.append(f"Se genero un archivo CSV del {nombre_almacen}")

    def leer_csv_basico(self):
        print("Lectura del CSV")
        try:
            with open(self.archivo, 'r') as f:
                lector = csv.reader(f)
                for fila in lector:
                    print(fila)
        except FileNotFoundError:
            print("Error: archivo no encontrado")

    def generar_excel(self, nombre_almacen="Almacen_principal"):
        
        try:

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Inventario_{nombre_almacen}"
            
            self.generar_csv_de_almacen(nombre_almacen)
            
            with open(self.archivo, 'r') as f:
                lector = csv.reader(f)
                datos = list(lector)
            
            for fila_num, fila_datos in enumerate(datos, 1):
                for col_num, valor in enumerate(fila_datos, 1):
                    # Convertimos los valores numéricos 
                    if fila_num > 1 and col_num in [3, 4, 5, 7]:  # Cantidad, Precio_Unitario, Precio_Total, Stock
                        try:
                            valor = float(valor) if col_num in [4, 5] else int(valor)
                        except:
                            pass
                    
                    ws.cell(row=fila_num, column=col_num, value=valor)
            
            self._formatear_encabezados(ws)
            
            self._ajustar_columnas_excel(ws)
            
            archivo_excel = f'inventario_{nombre_almacen}.xlsx'
            wb.save(archivo_excel)
            print(f"Archivo Excel creado: {archivo_excel}")
            lista_cambios.append(f"Se genero un archivo Excel del {nombre_almacen}")
            
        except Exception as e:
            print(f"Error al generar Excel: {e}")
    
    def _formatear_encabezados(self, ws):
        font_encabezado = Font(bold=True, color="FFFFFF")
        fill_encabezado = PatternFill(start_color="2F5F8F", end_color="2F5F8F", fill_type="solid")
        alineacion = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, 8): 
            cell = ws.cell(row=1, column=col)
            cell.font = font_encabezado
            cell.fill = fill_encabezado
            cell.alignment = alineacion
    
    def _ajustar_columnas_excel(self, ws):
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
#Metodos de paquetes y productos
def crear_producto(nombre:str, precio:float, peso:float, categoria:str, stock:int)->"Producto":
    mercancia=Producto(nombre, precio, peso, categoria, stock)
    catalogo_productos[nombre]=mercancia
    lista_cambios.append(f"Se agrego el producto {nombre} al catagolo de mercancias con un precio de {precio}")
    return mercancia

def crear_paquete(cantidad:int, contenido:"Producto"):
    nuevo_paquete=Paquete(cantidad, contenido)
    lista_cambios.append(f"Se armo el paquete {nuevo_paquete.id}")
    return nuevo_paquete

# Función nueva para exportar
def exportar_almacen_excel(nombre_almacen="Almacen_principal"):
    manejador = ManejadorCSV()
    manejador.generar_excel(nombre_almacen)

def exportar_almacen_csv(nombre_almacen="Almacen_principal"):
    manejador = ManejadorCSV()
    manejador.generar_csv_de_almacen(nombre_almacen)

# Menú básico para el sistema de almacén

def mostrar_menu():
    print("\n" + "="*40)
    print("    SISTEMA DE ALMACÉN v1.6.2")
    print("="*40)
    print("1. Crear nuevo producto")
    print("2. Crear paquete")
    print("3. Ver productos del catálogo")
    print("4. Ver contenido de almacén")
    print("5. Exportar almacén a Excel")
    print("6. Ver historial de cambios")
    print("7. Ver registros con fechas")
    print("8. Salir")
    print("="*40)

def menu_crear_producto():
    print("\n--- CREAR NUEVO PRODUCTO ---")
    nombre = input("Nombre del producto: ")
    
    try:
        precio = float(input("Precio: $"))
        peso = float(input("Peso (kg): "))
        stock = int(input("Stock inicial: "))
    except ValueError:
        print("Error: Debes introducir números válidos")
        return
    
    print("\nCategorías disponibles:")
    print("P - Paquete de prueba")
    print("F - Frágil") 
    print("X - No marcado")
    print("E - Eléctrico")
    print("VA - Forma de vida")
    
    categoria = input("Categoría (P/F/X/E/VA): ").upper()
    
    if categoria not in ["P", "F", "X", "E", "VA"]:
        print("Categoría no válida")
        return
    
    producto = crear_producto(nombre, precio, peso, categoria, stock)
    print(f"Producto '{nombre}' creado exitosamente")

def menu_crear_paquete():
    if not catalogo_productos:
        print("No hay productos en el catálogo. Crea un producto primero.")
        return
    
    print("\n--- CREAR PAQUETE ---")
    print("Productos disponibles:")
    for i, nombre in enumerate(catalogo_productos.keys(), 1):
        print(f"{i}. {nombre}")
    
    try:
        opcion = int(input("Selecciona el número del producto: ")) - 1
        nombres = list(catalogo_productos.keys())
        
        if 0 <= opcion < len(nombres):
            producto_elegido = catalogo_productos[nombres[opcion]]
            cantidad = int(input("Cantidad para el paquete: "))
            
            paquete = crear_paquete(cantidad, producto_elegido)
            
            print("\nAlmacenes disponibles:")
            for almacen in ALMACENES.keys():
                print(f"- {almacen}")
            
            almacen = input("¿En qué almacén guardar? (default: Almacen_principal): ").strip()
            if not almacen:
                almacen = "Almacen_principal"
            
            if almacen in ALMACENES:
                paquete.guardar_paquete(almacen)
                print(f"Paquete {paquete.id} creado y guardado en {almacen}")
            else:
                print("Almacén no válido")
        else:
            print("Opción no válida")
    except ValueError:
        print("Debes introducir un número válido")

def ver_catalogo():
    if not catalogo_productos:
        print("El catálogo está vacío")
        return
    
    print("\n--- CATÁLOGO DE PRODUCTOS ---")
    for nombre, producto in catalogo_productos.items():
        print(f"• {nombre}: ${producto.get_precio()} - Stock: {producto.stock} - Categoría: {producto.categoria}")
def ver_almacen():
    print("\nCONTENIDO DE ALMACENES")
    for nombre_almacen, paquetes in ALMACENES.items():
        print(f"\n{nombre_almacen}: {len(paquetes)} paquetes")
        if paquetes:
            for paquete in paquetes:
                print(f"  - {paquete.id}: {paquete.cantidad}x {paquete.contenido._nombre}")

def ver_historial_cambios():
    print("\n--- HISTORIAL DE CAMBIOS ---")
    if not lista_cambios:
        print("No hay cambios registrados aún")
    else:
        print(f"Total de cambios: {len(lista_cambios)}")
        for i, cambio in enumerate(lista_cambios, 1):
            print(f"{i}. {cambio}")

def ver_registros_completos():
    print("\n--- REGISTROS CON FECHAS ---")
    if not registros_a:
        print("No hay registros guardados aún")
        print("Los registros se crean automáticamente con las funciones")
        return
    
    print(f"Total de registros: {len(registros_a)}")
    for registro_id, registro in registros_a.items():
        fecha_simple = registro.fecha.strftime('%d/%m/%Y %H:%M')
        print(f"\n{registro_id} - {fecha_simple}")
        print(f"   Notas: {registro._notas_operador}")
        if registro.cambios:
            print("   Cambios:")
            for cambio in registro.cambios:
                print(f"     - {cambio}")
        else:
            print("   Sin cambios registrados")
    print("\n--- CONTENIDO DE ALMACENES ---")
    for nombre_almacen, paquetes in ALMACENES.items():
        print(f"\n{nombre_almacen}: {len(paquetes)} paquetes")
        if paquetes:
            for paquete in paquetes:
                print(f"  - {paquete.id}: {paquete.cantidad}x {paquete.contenido._nombre}")

def menu_exportar():
    print("\n--- EXPORTAR ALMACÉN ---")
    print("Almacenes disponibles:")
    for almacen in ALMACENES.keys():
        print(f"- {almacen}")
    
    almacen = input("¿Qué almacén exportar? (default: Almacen_principal): ").strip()
    if not almacen:
        almacen = "Almacen_principal"
    
    if almacen in ALMACENES:
        if ALMACENES[almacen]:
            exportar_almacen_excel(almacen)
        else:
            print(f"El almacén {almacen} está vacío")
    else:
        print("Almacén no válido")

def main():
    
    while True:
        mostrar_menu()
        
        try:
            opcion = input("Selecciona una opción (1-6): ").strip()
            
            if opcion == "1":
                menu_crear_producto()
            elif opcion == "2":
                menu_crear_paquete()
            elif opcion == "3":
                ver_catalogo()
            elif opcion == "4":
                ver_almacen()
            elif opcion == "5":
                menu_exportar()
            elif opcion == "6":
                ver_historial_cambios()
            elif opcion == "7":
                ver_registros_completos()
            elif opcion == "8":
                print("¡Gracias por usar el sistema de almacén!")
                break
            else:
                print("Opción no válida. Elige entre 1-8")
                
            input("\nPresiona Enter para continuar...")
            
        except KeyboardInterrupt:
            print("\n\n¡Hasta luego!")
            break

if __name__ == "__main__":
    main()

#prueba de productos y exportacion a excel
'''laptop = crear_producto("Laptop", 1500.0, 2.5, "E", 10)
mouse = crear_producto("Mouse inalámbrico", 35.0, 0.1, "E", 50)
vaso_cristal = crear_producto("Vaso de cristal", 8.0, 0.3, "F", 25)
laptop2 = crear_producto("Laptop", 1500.0, 2.5, "E", 10)
mouse2 = crear_producto("Mouse inalámbrico", 35.0, 0.1, "E", 50)
vaso_cristal2 = crear_producto("Vaso de cristal", 8.0, 0.3, "F", 25)
laptop3 = crear_producto("Laptop", 1500.0, 2.5, "E", 10)
mouse3 = crear_producto("Mouse inalámbrico", 35.0, 0.1, "E", 50)
vaso_cristal3 = crear_producto("Vaso de cristal", 8.0, 0.3, "F", 25)
laptop4 = crear_producto("Laptop", 1500.0, 2.5, "E", 10)
mouse4 = crear_producto("Mouse inalámbrico", 35.0, 0.1, "E", 50)
vaso_cristal4 = crear_producto("Vaso de cristal", 8.0, 0.3, "F", 25)
paquete1 = crear_paquete(2, laptop)
paquete2 = crear_paquete(10, mouse)
paquete3 = crear_paquete(5, vaso_cristal)
paquete4 = crear_paquete(2, laptop)
paquete5 = crear_paquete(10, mouse)
paquete6 = crear_paquete(5, vaso_cristal)
paquete7 = crear_paquete(2, laptop)
paquete8 = crear_paquete(10, mouse)
paquete9 = crear_paquete(5, vaso_cristal)
paquete1.guardar_paquete("Almacen_principal")
paquete2.guardar_paquete("Almacen_principal")
paquete3.guardar_paquete("Almacen_principal")
paquete4.guardar_paquete("Almacen_principal")
paquete5.guardar_paquete("Almacen_principal")
paquete6.guardar_paquete("Almacen_principal")
paquete7.guardar_paquete("Almacen_principal")
paquete8.guardar_paquete("Almacen_principal")
paquete9.guardar_paquete("Almacen_principal")
#Prueba de exportación
print("Exportando almacen principal a Excel")
exportar_almacen_excel("Almacen_principal")'''

#Prueba de impresion-------------------------------------
#print("Deseas obtener la guia de ids de paquetes actual?")
#solicitud_operador=input("Y/N:")
#v = solicitud_operador.lower() == "y" #  tambien sirve v = True if solicitud_operador in ("Y", "y") else False

#if v == True:
 # for id , ref in guia_de_ids.items():
 #   print(f"{id}:{ref}\n")

#Pruebas Registro----------------------------------------
#registro1=Registro("Hola wenas",[])
#registro1.mostrar_registro()
#print(f"{registro1.id} - {registro1.N_registro}\n")
#registro2=Registro("Que mas!",["NO CAMBIA NADA"],"A")
#registro2.mostrar_registro()
#print(f"{registro2.id} - {registro2.N_registro}\n")
#registro3=Registro("Nota nueva!",[], "X")

#new_reg= crear_registro(input("Agrega tus notas aqui:"),"A")
#new_reg.mostrar_registro()
#r1=input("Deseas descargar algo? y/n:")
#if r1 == "y":
 #   descargar()

#r=input("Deseas descargar algo? y/n:")
#if r == "y":
#    cargar()
#print(registros_a)
#---------------------------------------------------------