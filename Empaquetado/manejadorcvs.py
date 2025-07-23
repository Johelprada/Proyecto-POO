from openpyxl.styles import Font, Alignment, PatternFill
from paquetes import lista_cambios,catalogo_productos, ALMACENES, Paquete, Producto

import csv
import openpyxl
class ManejadorCSV:
    def __init__(self):
        self.archivo = 'productos_almacen.csv'
    
    def generar_csv_de_almacen(self, nombre_almacen="Almacen_principal"):
        datos = [['ID_Paquete', 'Contenido', 'Cantidad', 'Precio_Unitario', 'Precio_Total', 'Categoria', 'Stock']]
        
        if nombre_almacen in ALMACENES:
            for paquete in ALMACENES[nombre_almacen]:
                fila = [
                    paquete.id,
                    paquete.contenido._nombre,
                    paquete.cantidad,
                    paquete.contenido.get_precio(),
                    paquete._precio,
                    paquete.contenido.categoria,
                    paquete.contenido.stock
                ]
                datos.append(fila)
        
        with open(self.archivo, 'w', newline='') as archivo_csv:
            escritor = csv.writer(archivo_csv)
            escritor.writerows(datos)
        
        print(f"Archivo CSV creado: {self.archivo}")
        lista_cambios.append(f"Se genero un archivo CSV del {nombre_almacen}")

    def leer_csv_basico(self):
        print("Lectura del CSV")
        try:
            with open(self.archivo, 'r') as f:
                lector = csv.reader(f)
                for fila in lector:
                    print(fila)
        except FileNotFoundError:
            print("Error: archivo no encontrado")

    def generar_excel(self, nombre_almacen="Almacen_principal"):
        
        try:

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Inventario_{nombre_almacen}"
            
            self.generar_csv_de_almacen(nombre_almacen)
            
            with open(self.archivo, 'r') as f:
                lector = csv.reader(f)
                datos = list(lector)
            
            for fila_num, fila_datos in enumerate(datos, 1):
                for col_num, valor in enumerate(fila_datos, 1):
                    # Convertimos los valores numéricos 
                    if fila_num > 1 and col_num in [3, 4, 5, 7]:  # Cantidad, Precio_Unitario, Precio_Total, Stock
                        try:
                            valor = float(valor) if col_num in [4, 5] else int(valor)
                        except:
                            pass
                    
                    ws.cell(row=fila_num, column=col_num, value=valor)
            
            self._formatear_encabezados(ws)
            
            self._ajustar_columnas_excel(ws)
            
            archivo_excel = f'inventario_{nombre_almacen}.xlsx'
            wb.save(archivo_excel)
            print(f"Archivo Excel creado: {archivo_excel}")
            lista_cambios.append(f"Se genero un archivo Excel del {nombre_almacen}")
            
        except Exception as e:
            print(f"Error al generar Excel: {e}")
    
    def _formatear_encabezados(self, ws):
        font_encabezado = Font(bold=True, color="FFFFFF")
        fill_encabezado = PatternFill(start_color="2F5F8F", end_color="2F5F8F", fill_type="solid")
        alineacion = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, 8): 
            cell = ws.cell(row=1, column=col)
            cell.font = font_encabezado
            cell.fill = fill_encabezado
            cell.alignment = alineacion
    
    def _ajustar_columnas_excel(self, ws):
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
#Metodos de paquetes y productos
def crear_producto(nombre:str, precio:float, peso:float, categoria:str, stock:int)->"Producto":
    mercancia=Producto(nombre, precio, peso, categoria, stock)
    catalogo_productos[nombre]=mercancia
    lista_cambios.append(f"Se agrego el producto {nombre} al catagolo de mercancias con un precio de {precio}")
    return mercancia

def crear_paquete(cantidad:int, contenido:"Producto"):
    nuevo_paquete=Paquete(cantidad, contenido)
    lista_cambios.append(f"Se armo el paquete {nuevo_paquete.id}")
    return nuevo_paquete

# Función nueva para exportar
def exportar_almacen_excel(nombre_almacen="Almacen_principal"):
    manejador = ManejadorCSV()
    manejador.generar_excel(nombre_almacen)

def exportar_almacen_csv(nombre_almacen="Almacen_principal"):
    manejador = ManejadorCSV()
    manejador.generar_csv_de_almacen(nombre_almacen)