import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

#creamos una clase producto con la cual vamos a poder generar los datos a los cuales queremos enlistar
class Producto:
    def __init__(self, id, nombre, precio, categoria, stock):
        self.id = id
        self.nombre = nombre
        self.precio = precio
        self.categoria = categoria
        self.stock = stock

#aqui se maneja el csv (tengo que estudiarla mas, no puedo solo seguir tutoriales)
class ManejadorCSV:
    def __init__(self):
        self.archivo = 'productos.csv'
    
    def crear_csv(self):
  # Aqui creamos los datos de prueba (se espera que estos datos al final sean los de la lista global osea el almacen principal)
        datos = [
            ['ID', 'Nombre', 'Precio', 'Categoria', 'Stock'],
            [1, 'computador', 850.99, 'Electronica', 15],
            [2, 'Mouse', 25.50, 'Accesorios', 50],
            [3, 'Teclado Mecanico', 89.99, 'Accesorios', 30],
            [4, 'Monitor 24"', 299.99, 'Electronica', 8],
            [5, 'Auriculares', 45.00, 'Audio', 25],
            [6, 'Teclado', 80.00, 'Electronica', 14]
        ]
        
        with open(self.archivo, 'w', newline='') as archivo_csv:
            escritor = csv.writer(archivo_csv)
            escritor.writerows(datos)
        
        print(f"Archivo {self.archivo} creado")

    def leer_csv_basico(self):
        print("--- Lectura basica con CSV ---")
        try:
            with open(self.archivo, 'r') as f:
                lector = csv.reader(f)
                for fila in lector:
                    print(fila)
        except FileNotFoundError:
            print("Error: archivo no encontrado")

#definimos los filtros que se van a poner, aqui tome el atributo precio, pero en el proyecto final tmb podriamos usar cosas como urgente, vivo, etc.
    def probar_filtros(self):
        print("\n--- Productos caros (>$50) ---")
        try:
            with open(self.archivo, 'r') as f:
                lector = csv.DictReader(f)
                for fila in lector:
                    if float(fila['Precio']) > 50:
                        print(f"{fila['Nombre']} - ${fila['Precio']}")
        except Exception as e:
            print(f"Error: {e}")
        
        print("\n--- Por categoria ---")
        try:
            with open(self.archivo, 'r') as f:
                lector = csv.DictReader(f)
                categorias = {}
                for fila in lector:
                    categoria = fila['Categoria']
                    precio = float(fila['Precio'])
                    
                    if categoria not in categorias:
                        categorias[categoria] = []
                    categorias[categoria].append(precio)
                
                for categoria, precios in categorias.items():
                    promedio = sum(precios) / len(precios)
                    print(f"{categoria}: ${promedio:.2f}")
        except Exception as e:
            print(f"Error: {e}")

    #con esto podemos generar el excel y hacer que se vea bonito
    #la mayor parte de esto esta hecho a base de tutoriales, no me funen si se ve muy basico
    def generar_excel(self):

        print("\n--- Generando archivo Excel ---")
        
        try:
            # Creamos un libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inventario de Productos"
            
            # Leeemos los datos del CSV existente
            with open(self.archivo, 'r') as f:
                lector = csv.reader(f)
                datos = list(lector)
            
            # Escribimos los datos en Excel
            for fila_num, fila_datos in enumerate(datos, 1):
                for col_num, valor in enumerate(fila_datos, 1):
                    # Convertir valores numéricos
                    if fila_num > 1 and col_num in [1, 3, 5]:  # ID, Precio, Stock
                        try:
                            valor = float(valor) if col_num == 3 else int(valor)
                        except:
                            pass
                    
                    ws.cell(row=fila_num, column=col_num, value=valor)
            
            # Aplicamos formato a encabezados
            self._formatear_encabezados(ws)
            
            # Ajustamos columnas
            self._ajustar_columnas_excel(ws)
            
            # Guardar el archivo
            archivo_excel = 'productos_inventario.xlsx'
            wb.save(archivo_excel)
            print(f"Archivo Excel creado: {archivo_excel} en la misma caroeta del archivo .py") #luego tengo que ver como hacer para que me lo genere en otra parte
            
        except Exception as e:
            print(f"Error al generar Excel: {e}")
    
    def _formatear_encabezados(self, ws):

        # Estilo para encabezados
        font_encabezado = Font(bold=True, color="FFFFFF")
        fill_encabezado = PatternFill(start_color="2F5F8F", end_color="2F5F8F", fill_type="solid")
        alineacion = Alignment(horizontal="center", vertical="center")
        
        # Esto de momento creo que solo sirve para el ejemplo, luego tengo qeu cambiarlo paraa que sirva para todos los paquetes q generemos
        for col in range(1, 6):  # 5 columnas
            cell = ws.cell(row=1, column=col)
            cell.font = font_encabezado
            cell.fill = fill_encabezado
            cell.alignment = alineacion
    
    def _ajustar_columnas_excel(self, ws):
        """Ajusta automáticamente el ancho de las columnas"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

def main():
    print(" PRUEBA DE CSV ")
    
    manejador = ManejadorCSV()
    
    # Creamos el CSV 
    manejador.crear_csv()
    
    # Leemos de forma basica 
    manejador.leer_csv_basico()
    
    # Probar algunos filtros 
    manejador.probar_filtros()
    
    #generamos el excel basado en el csv
    manejador.generar_excel()

if __name__ == "__main__":
    main()
